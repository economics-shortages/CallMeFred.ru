import openpyxl
import dateparser

from parser.core.db.connections import Connections
from parser.core.parser.parser import Parser
from parser.core.models.metrics import Metrics
from pathlib import Path

class Bank_Sector_Loans_Db():
    TARGET_URL = 'https://www.cbr.ru/vfs/statistics/pdko/int_rat/loans_nonfin.xlsx'
    TARGET_FILE = '/app_data/int_rat_loans_nonfin.xlsx'

    cn: Connections

    parser = None
    sheet_name_to_currency_map = {
        "ставки_руб": "RUB",
        "ставки_долл.США": "USD",
        "ставки_евро": "EUR",
        "cтруктура_руб": "RUB",
        "структура_ долл.США": "USD",
        "структура_евро": "EUR"
    }

    def __init__(self) -> None:
        self.cn = Connections()

    def get_metric(self):
        pass
    
    def prepare_value(self, value) -> int:
        try:
            val = int(value * 100)
        except ValueError:
            val = 0
        return val

    def start(self):
        self.parser = Parser().get_parser(browserless=True)
        self.parser.init_url(self.TARGET_URL)
        self.parser.save(self.TARGET_FILE)
        self.upload_to_db()

    def sheet_name_to_currency(self, sheet_name) -> dict:
        if sheet_name not in self.sheet_name_to_currency_map:
            raise Exception('Can\'t recognize sheet')
        currency_name = self.sheet_name_to_currency_map[sheet_name]
        query = 'SELECT id FROM currency WHERE code = %s;'
        return self.cn.query_single(query, (currency_name))

    def init_metric(self, title: str, currency: dict) -> Metrics:
        try:
            m = Metrics().load_by_fields({
                'currency': currency['id'],
                's_title': title
            })

        except:
            m = Metrics().set_data({
                'currency': currency['id'],
                's_title': title
            })
            m.save()

        return m
    
    def add_metric_value(self, metric, value, date):
        new_query = 'INSERT INTO `metric` (`id`, `metric`, `date_described`, `date_parsed`, `value`, `obsolete`) VALUES(NULL, %s, %s, NOW(), %s, 0);'
        last_id = self.cn.query_insert(new_query, (
            metric.id,
            date,
            value
        ))

    def upload_to_db(self):
        xlsx_file = Path(self.TARGET_FILE)
        wb = openpyxl.load_workbook(xlsx_file, read_only=True) 

        for sheet in wb:
            self.parse_sheet(sheet)

    def parse_sheet(self, sheet: openpyxl.worksheet):
        currency = self.sheet_name_to_currency(sheet.title)
        metric_head = []
        nm_rows = sheet.max_row
        for row in sheet.iter_rows(min_row=1, max_row=2):
            for cell in row:
                if cell.value is not None:
                    metric_head.append(cell.value)
        metric_head = ' / '.join(metric_head)

        root_metric = self.init_metric(metric_head, currency)
        
        big_metrics = []
        small_metrics = []
        cols_small = []

        for row in sheet.iter_rows(min_row=4, max_row=4):
            for cell in row:
                if cell.value is not None:
                    big_metric = self.init_metric(cell.value, currency)
                    big_metric.set_parent(root_metric.id)
                    big_metrics.append(big_metric)
        
        for row in sheet.iter_rows(min_row=5, max_row=5):
            for cell in row:
                if cell.value is not None:
                    cols_small.append(cell.value)

        big_col_size = int(len(cols_small)/2)
        
        small_metrics = []
        for i in range(0, len(cols_small)):
            cl_dir = int(i < big_col_size)
            
            pm = big_metrics[cl_dir]
            pm_title = pm.get('s_title')
            new_col_nm = pm_title + ' / ' + cols_small[i]
            sm = self.init_metric(new_col_nm, currency)
            sm.set_parent(pm.id)
            small_metrics.append(sm)
        
        del cols_small

        for row in sheet.iter_rows(min_row=6, max_row=nm_rows-2):
            cel_vals = []
            for cell in row:
                if cell.value is not None:
                    cel_vals.append(cell.value)
            dt = dateparser.parse(str(cel_vals[0]), date_formats=['%B %Y'])
            cols_vals = cel_vals[1:]

            nm_cols = len(cols_vals)
            for i in range(0, nm_cols):
                metric = small_metrics[i]
                value = self.prepare_value(cols_vals[i])
                self.add_metric_value(metric, value, dt)